import speech_recognition as sr
import pyttsx3
import pywhatkit
import datetime
import wikipedia
import pyjokes

listener = sr.Recognizer()
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)


def talk(text):
    engine.say(text)
    engine.runAndWait()


def take_command():
    try:
        with sr.Microphone() as source:
            print('listening...')
            voice = listener.listen(source)
            command = listener.recognize_google(voice)
            command = command.lower()
            if 'alexa' in command:
                command = command.replace('alexa', '')
                print(command)
    except:
        pass
    return command


def run_alexa():
    command = take_command()
    print(command)
    if 'play' in command:
        song = command.replace('play', '')
        talk('playing ' + song)
        pywhatkit.playonyt(song)
    elif 'time' in command:
        time = datetime.datetime.now().strftime('%I:%M %p')
        talk('Current time is ' + time)
    elif 'who the heck is' in command:
        person = command.replace('who the heck is', '')
        info = wikipedia.summary(person, 1)
        print(info)
        talk(info)
    elif 'date' in command:
        talk('sorry, I am busy right now')
    elif 'are you single' in command:
        talk('I am in a relationship with golden ')
    elif 'joke' in command:
        talk(pyjokes.get_joke())
    elif ' do you know shivali saha' in command:
        talk('yes , she is my upcoming owner')
    elif 'marry me' in command:
        talk ('sorry, i am not human being i am a machine')
    else:
        talk('Please say the command again.')


while True:
    run_alexa()




import pandas as pd
import openpyxl
from openpyxl.styles import Side, Border

size = len(aggregated_df_yearly3.columns)

# Create an ExcelWriter object
with pd.ExcelWriter('Split Yearly Summary.xlsx', engine='openpyxl') as writer:
    writer.book = openpyxl.load_workbook('Split Yearly Summary.xlsx')
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    # Write Re_adjusted_contribution_new_yearlyI with heading and blue color
    k = 0
    for j in model_list:
        globals()[f'Re_adjusted_contribution_new_yearly{j}'].to_excel(writer, sheet_name='yearly summary', startrow=1, startcol=k*(size+2), index=True)
        worksheet = writer.sheets['yearly summary']
        worksheet.cell(row=1, column=k*(size+2)).value = f'Yearly Aggregation of Model {j}'
        # Add borders to each cell within the table
        for table in worksheet.tables.values():
            start_row = table.ref.topLeftCell.row
            end_row = table.ref.bottomRightCell.row
            start_column = table.ref.topLeftCell.column
            end_column = table.ref.bottomRightCell.column
            for row in range(start_row, end_row + 1):
                for col in range(start_column, end_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    border = Border(top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'),
                                    left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'))
                    cell.border = border
        k += 1

    # Write aggregated_df_yearly3 with heading and blue color
    aggregated_df_yearly3.to_excel(writer, sheet_name='yearly summary', startrow=1, startcol=k*(size+2), index=True)
    worksheet = writer.sheets['yearly summary']
    worksheet.cell(row=1, column=k*(size+2)).value = 'Yearly Agg of All Models'
    # Add borders to each cell within the table
    for table in worksheet.tables.values():
        start_row = table.ref.topLeftCell.row
        end_row = table.ref.bottomRightCell.row
        start_column = table.ref.topLeftCell.column
        end_column = table.ref.bottomRightCell.column
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                border = Border(top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000'),
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'))
                cell.border = border

# Save the modified workbook
writer.save()